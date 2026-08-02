# -*- coding: utf-8 -*-
"""Spreadsheet templates, parsers, previews, and exports for company catalogs."""

import csv
import datetime
import hashlib
import io
import json
import re
import secrets
import shutil
import time
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from admin_common import ApiError, now_iso
from aicm_io import data_lock, read_json, write_json_atomic
from ticker_meta import normalize_ticker, resolve_ticker
from watchlist_loader import (
    load_active_universe,
    load_user_watchlist,
    rebuild_active_universe,
    write_user_watchlist,
)


XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
IMPORT_COLUMNS = (
    "ticker",
    "name",
    "country",
    "city",
    "lat",
    "lon",
    "seg",
    "chain",
    "chain_key",
    "enabled",
)
EXPORT_COLUMNS = (
    "ticker",
    "name",
    "country",
    "city",
    "lat",
    "lon",
    "seg",
    "chain",
    "chain_key",
    "source",
    "enabled",
    "latest_price",
    "change_pct",
    "quote_time",
    "quote_source",
)
MAX_UPLOAD_BYTES = 10 * 1024 * 1024
MAX_IMPORT_ROWS = 10000
PREVIEW_TTL_SECONDS = 30 * 60
REPLACE_CONFIRMATION = "REPLACE ACTIVE CATALOG"


def _style_header(sheet, row=1):
    fill = PatternFill("solid", fgColor="C7A65B")
    for cell in sheet[row]:
        cell.fill = fill
        cell.font = Font(color="111318", bold=True)
        cell.alignment = Alignment(vertical="center")
    sheet.row_dimensions[row].height = 24


def _set_widths(sheet, widths):
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width


def _workbook_bytes(workbook):
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def template_bytes(variant):
    variant = str(variant or "blank").strip().lower()
    if variant not in {"blank", "example"}:
        raise ApiError(400, "invalid_template_variant", "Template variant must be blank or example")

    workbook = Workbook()
    instructions = workbook.active
    instructions.title = "Instructions"
    companies = workbook.create_sheet("Companies")
    reference = workbook.create_sheet("Field Reference")

    instruction_rows = [
        ("AI 算力链上市公司名录导入模板", "请勿修改工作表名称或 Companies 表头。"),
        ("操作流程", "填写 Companies → 上传预检 → 查看差异 → 确认导入。"),
        ("安全规则", "空白单元格不会覆盖现有值；未出现在文件中的公司保持不变。"),
        ("股票代码示例", "NVDA、0700.HK、600519.SS、2330.TW、005930.KS。"),
        ("enabled", "TRUE 表示启用，FALSE 表示停用；留空表示保持当前状态。"),
        ("文件限制", "仅支持 .xlsx 和 UTF-8 CSV，最多 10,000 行、10 MB。"),
    ]
    for row in instruction_rows:
        instructions.append(row)
    instructions["A1"].font = Font(size=16, bold=True, color="C7A65B")
    instructions.freeze_panes = "A2"
    _set_widths(instructions, {"A": 22, "B": 88})

    companies.append(IMPORT_COLUMNS)
    if variant == "example":
        companies.append(("NVDA", "NVIDIA 英伟达", "美国", "Santa Clara", 37.35, -121.95, "GPU/AI芯片", "AI芯片", "chip", True))
        companies.append(("600519.SS", "贵州茅台", "中国大陆", "Guiyang", 26.65, 106.63, "示例分类", "其他", "other", True))
    _style_header(companies)
    companies.freeze_panes = "A2"
    companies.auto_filter.ref = "A1:J1"
    _set_widths(
        companies,
        {"A": 18, "B": 28, "C": 16, "D": 20, "E": 12, "F": 12, "G": 24, "H": 20, "I": 18, "J": 12},
    )
    enabled_validation = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)
    companies.add_data_validation(enabled_validation)
    enabled_validation.add("J2:J10001")

    reference.append(("field", "required", "type", "example", "validation"))
    field_rows = [
        ("ticker", "是", "文本", "NVDA", "Yahoo Finance 支持的股票代码"),
        ("name", "否", "文本", "NVIDIA 英伟达", "空白时保留原值或自动解析"),
        ("country", "否", "文本", "美国", "国家或地区"),
        ("city", "否", "文本", "Santa Clara", "总部或主要运营城市"),
        ("lat", "否", "数字", "37.35", "-90 至 90"),
        ("lon", "否", "数字", "-121.95", "-180 至 180"),
        ("seg", "否", "文本", "GPU/AI芯片", "产业细分"),
        ("chain", "否", "文本", "AI芯片", "产业链名称"),
        ("chain_key", "否", "文本", "chip", "稳定的产业链分类键"),
        ("enabled", "否", "布尔", "TRUE", "TRUE、FALSE 或空白"),
    ]
    for row in field_rows:
        reference.append(row)
    _style_header(reference)
    reference.freeze_panes = "A2"
    _set_widths(reference, {"A": 18, "B": 12, "C": 14, "D": 24, "E": 48})

    suffix = "示例" if variant == "example" else "空白"
    return _workbook_bytes(workbook), f"AI算力链上市公司导入模板_{suffix}.xlsx", XLSX_MIME


def _normalize_headers(headers):
    normalized = []
    seen = set()
    for raw in headers or []:
        name = str(raw or "").strip().lower()
        if not name:
            raise ApiError(400, "invalid_headers", "Import contains a blank column header")
        if name in seen:
            raise ApiError(400, "duplicate_header", f"Duplicate column header: {name}")
        if name not in IMPORT_COLUMNS:
            raise ApiError(400, "unsupported_column", f"Unsupported import column: {name}")
        normalized.append(name)
        seen.add(name)
    if "ticker" not in seen:
        raise ApiError(400, "missing_ticker_column", "Import requires a ticker column")
    return normalized


def _clean_cell(value):
    if isinstance(value, str):
        return value.strip()
    return value


def _rows_from_matrix(headers, values):
    headers = _normalize_headers(headers)
    rows = []
    for row_number, raw_values in values:
        raw_values = list(raw_values)
        record = {
            header: _clean_cell(raw_values[index] if index < len(raw_values) else None)
            for index, header in enumerate(headers)
        }
        if all(value in (None, "") for value in record.values()):
            continue
        record["_row"] = row_number
        rows.append(record)
        if len(rows) > MAX_IMPORT_ROWS:
            raise ApiError(413, "too_many_rows", f"Import exceeds {MAX_IMPORT_ROWS} rows")
    return rows


def _parse_csv(body):
    try:
        text = body.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise ApiError(400, "invalid_encoding", "CSV must use UTF-8 encoding") from exc
    reader = csv.reader(io.StringIO(text))
    try:
        headers = next(reader)
    except StopIteration as exc:
        raise ApiError(400, "empty_file", "Import file is empty") from exc
    return _rows_from_matrix(headers, ((index, row) for index, row in enumerate(reader, start=2)))


def _parse_xlsx(body):
    try:
        workbook = load_workbook(io.BytesIO(body), read_only=True, data_only=True)
    except Exception as exc:
        raise ApiError(400, "invalid_workbook", "XLSX file could not be opened") from exc
    if "Companies" not in workbook.sheetnames:
        raise ApiError(400, "missing_companies_sheet", "XLSX requires a Companies sheet")
    sheet = workbook["Companies"]
    iterator = sheet.iter_rows(values_only=True)
    try:
        headers = next(iterator)
    except StopIteration as exc:
        raise ApiError(400, "empty_file", "Companies sheet is empty") from exc
    return _rows_from_matrix(headers, ((index, row) for index, row in enumerate(iterator, start=2)))


def parse_upload(filename, body):
    if not isinstance(body, (bytes, bytearray)):
        raise ApiError(400, "invalid_upload", "Upload body must be binary")
    if len(body) > MAX_UPLOAD_BYTES:
        raise ApiError(413, "payload_too_large", "Upload exceeds 10 MB")
    suffix = Path(str(filename or "")).suffix.lower()
    if suffix == ".csv":
        return _parse_csv(bytes(body))
    if suffix == ".xlsx":
        return _parse_xlsx(bytes(body))
    raise ApiError(400, "unsupported_file_type", "Only .xlsx and .csv files are supported")


def _scope_rows(base, scope, filters=None):
    base = Path(base)
    scope = str(scope or "active").lower()
    if scope == "active":
        rows = read_json(base / "data" / "active_universe.json", {}).get("companies", [])
    elif scope == "base":
        rows = read_json(base / "data" / "watchlist.json", {}).get("companies", [])
    elif scope == "user":
        user = read_json(base / "data" / "user_watchlist.json", {})
        rows = [dict(row) for row in (user.get("add") or {}).values()]
        rows.extend({"t": ticker, "enabled": False, "source": "removed"} for ticker in user.get("remove", []))
    elif scope == "filtered":
        rows = list((filters or {}).get("rows") or [])
    else:
        raise ApiError(400, "invalid_export_scope", "Unsupported export scope")
    return [dict(row) for row in rows]


def _quote_items(base):
    payload = read_json(Path(base) / "data" / "live.json", {})
    items = payload.get("items", {}) if isinstance(payload, dict) else {}
    if isinstance(items, dict):
        return items
    return {str(row.get("t") or "").upper(): row for row in items if row.get("t")} if isinstance(items, list) else {}


def _export_rows(base, rows, include_quotes):
    quotes = _quote_items(base) if include_quotes else {}
    output = []
    for source in rows:
        ticker = str(source.get("t") or source.get("ticker") or "").upper()
        quote = quotes.get(ticker) or {}
        output.append({
            "ticker": ticker,
            "name": source.get("name"),
            "country": source.get("country"),
            "city": source.get("city"),
            "lat": source.get("lat"),
            "lon": source.get("lon"),
            "seg": source.get("seg"),
            "chain": source.get("chain"),
            "chain_key": source.get("chain_key"),
            "source": source.get("source"),
            "enabled": source.get("enabled", True),
            "latest_price": quote.get("p") if include_quotes else None,
            "change_pct": quote.get("chg") if include_quotes else None,
            "quote_time": quote.get("market_time") if include_quotes else None,
            "quote_source": quote.get("source") if include_quotes else None,
        })
    return sorted(output, key=lambda row: row["ticker"])


def _xlsx_export(rows, metadata):
    workbook = Workbook()
    companies = workbook.active
    companies.title = "Companies"
    companies.append(EXPORT_COLUMNS)
    for row in rows:
        companies.append(tuple(row.get(column) for column in EXPORT_COLUMNS))
    _style_header(companies)
    companies.freeze_panes = "A2"
    companies.auto_filter.ref = f"A1:O{max(1, len(rows) + 1)}"
    _set_widths(
        companies,
        {"A": 18, "B": 30, "C": 16, "D": 20, "E": 12, "F": 12, "G": 24, "H": 20, "I": 18, "J": 16, "K": 12, "L": 16, "M": 14, "N": 18, "O": 18},
    )
    metadata_sheet = workbook.create_sheet("Metadata")
    metadata_sheet.append(("field", "value"))
    for key, value in metadata.items():
        metadata_sheet.append((key, json.dumps(value, ensure_ascii=False) if isinstance(value, (dict, list)) else value))
    _style_header(metadata_sheet)
    _set_widths(metadata_sheet, {"A": 24, "B": 80})
    return _workbook_bytes(workbook)


def export_bytes(base, format, scope, filters=None, include_quotes=False):
    format = str(format or "xlsx").lower()
    scope = str(scope or "active").lower()
    rows = _export_rows(base, _scope_rows(base, scope, filters), bool(include_quotes))
    metadata = {
        "schema_version": 1,
        "generated_at": now_iso(),
        "scope": scope,
        "include_quotes": bool(include_quotes),
        "row_count": len(rows),
        "quote_note": "Near-real-time where available; delayed by provider/exchange",
    }
    stamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    stem = f"aicm_companies_{scope}_{stamp}"
    if format == "json":
        body = json.dumps({**metadata, "companies": rows}, ensure_ascii=False, indent=2).encode("utf-8")
        return body, stem + ".json", "application/json; charset=utf-8"
    if format == "csv":
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=EXPORT_COLUMNS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)
        return output.getvalue().encode("utf-8-sig"), stem + ".csv", "text/csv; charset=utf-8"
    if format == "xlsx":
        return _xlsx_export(rows, metadata), stem + ".xlsx", XLSX_MIME
    raise ApiError(400, "invalid_export_format", "Export format must be xlsx, csv, or json")


def _catalog_hash(base):
    digest = hashlib.sha256()
    for name in ("watchlist.json", "user_watchlist.json", "active_universe.json"):
        path = Path(base) / "data" / name
        digest.update(name.encode("utf-8"))
        digest.update(path.read_bytes() if path.exists() else b"<missing>")
    return digest.hexdigest()


def _preview_dir(base):
    path = Path(base) / "data" / "admin_previews"
    path.mkdir(parents=True, exist_ok=True)
    return path


def _cleanup_previews(base):
    now = time.time()
    for path in _preview_dir(base).glob("*.json"):
        payload = read_json(path, {})
        if float(payload.get("expires_epoch") or 0) <= now:
            path.unlink(missing_ok=True)


def _catalog_index(base):
    base = Path(base)
    watchlist = read_json(base / "data" / "watchlist.json", {})
    user = load_user_watchlist(base)
    active = load_active_universe(base)
    active_tickers = {str(row.get("t") or "").upper() for row in active.get("companies", [])}
    rows = {}
    for row in watchlist.get("companies", []):
        ticker = str(row.get("t") or "").upper()
        if ticker:
            rows[ticker] = {**row, "t": ticker, "source": "base", "enabled": ticker in active_tickers}
    for ticker, row in user.get("add", {}).items():
        ticker = str(ticker).upper()
        rows[ticker] = {
            **rows.get(ticker, {}),
            **row,
            "t": ticker,
            "source": row.get("source") or "user",
            "enabled": ticker in active_tickers,
        }
    for row in active.get("companies", []):
        ticker = str(row.get("t") or "").upper()
        if ticker:
            rows[ticker] = {**rows.get(ticker, {}), **row, "t": ticker, "enabled": True}
    return rows


def _optional_bool(value):
    if value in (None, ""):
        return None
    if isinstance(value, bool):
        return value
    text = str(value).strip().lower()
    if text in {"true", "1", "yes", "y", "是", "启用"}:
        return True
    if text in {"false", "0", "no", "n", "否", "停用"}:
        return False
    raise ValueError("enabled must be TRUE, FALSE, or blank")


def _preview_values(raw):
    errors = []
    values = {}
    try:
        ticker = normalize_ticker(raw.get("ticker"))
        values["ticker"] = ticker
    except Exception as exc:
        ticker = str(raw.get("ticker") or "").strip().upper()
        errors.append(str(exc))
        values["ticker"] = ticker

    for field in IMPORT_COLUMNS:
        if field in {"ticker", "enabled"}:
            continue
        value = raw.get(field)
        if value not in (None, ""):
            values[field] = value.strip() if isinstance(value, str) else value
    for field, lower, upper in (("lat", -90, 90), ("lon", -180, 180)):
        if field in values:
            try:
                number = float(values[field])
                if not lower <= number <= upper:
                    raise ValueError(f"{field} must be between {lower} and {upper}")
                values[field] = number
            except (TypeError, ValueError) as exc:
                errors.append(str(exc))
    try:
        values["enabled"] = _optional_bool(raw.get("enabled"))
    except ValueError as exc:
        values["enabled"] = None
        errors.append(str(exc))
    return values, errors


def _same_value(left, right):
    if isinstance(left, (int, float)) or isinstance(right, (int, float)):
        try:
            return float(left) == float(right)
        except (TypeError, ValueError):
            return False
    return str(left or "").strip() == str(right or "").strip()


def create_preview(base, rows, filename, actor="admin"):
    base = Path(base)
    _cleanup_previews(base)
    current = _catalog_index(base)
    seen = set()
    classified = []
    summary = {key: 0 for key in ("new", "update", "unchanged", "invalid", "duplicate", "would_disable", "conflict")}

    for raw in rows:
        values, errors = _preview_values(raw)
        ticker = values.get("ticker") or ""
        if ticker in seen and ticker:
            classification = "duplicate"
            errors.append("duplicate ticker in upload")
        elif errors:
            classification = "invalid"
        else:
            seen.add(ticker)
            existing = current.get(ticker)
            enabled = values.get("enabled")
            if existing and enabled is False and existing.get("enabled"):
                classification = "would_disable"
            elif not existing:
                classification = "new"
            else:
                changes = {
                    field: value for field, value in values.items()
                    if field not in {"ticker", "enabled"}
                    and not _same_value(existing.get(field), value)
                }
                enable_change = enabled is not None and bool(existing.get("enabled")) != enabled
                classification = "update" if changes or enable_change else "unchanged"
        summary[classification] += 1
        classified.append({
            "row": int(raw.get("_row") or 0),
            "ticker": ticker,
            "classification": classification,
            "valid": classification not in {"invalid", "duplicate", "conflict"},
            "values": values,
            "errors": errors,
        })

    preview_id = secrets.token_hex(16)
    created_epoch = time.time()
    preview = {
        "preview_id": preview_id,
        "filename": str(filename or "upload"),
        "actor": str(actor),
        "created_at": now_iso(),
        "created_epoch": created_epoch,
        "expires_epoch": created_epoch + PREVIEW_TTL_SECONDS,
        "catalog_hash": _catalog_hash(base),
        "summary": summary,
        "rows": classified,
    }
    write_json_atomic(_preview_dir(base) / f"{preview_id}.json", preview)
    return preview


def load_preview(base, preview_id):
    preview_id = str(preview_id or "").strip().lower()
    if not re.fullmatch(r"[0-9a-f]{32}", preview_id):
        raise ApiError(400, "invalid_preview_id", "Invalid preview identifier")
    path = _preview_dir(base) / f"{preview_id}.json"
    preview = read_json(path, {})
    if not preview:
        _cleanup_previews(base)
        raise ApiError(404, "preview_not_found", "Import preview was not found or has expired")
    if float(preview.get("expires_epoch") or 0) <= time.time():
        path.unlink(missing_ok=True)
        _cleanup_previews(base)
        raise ApiError(410, "preview_expired", "Import preview has expired")
    _cleanup_previews(base)
    return preview


def _default_mapping(default_companies):
    return default_companies() if callable(default_companies) else dict(default_companies or {})


def _resolved_import_rows(base, preview):
    current = _catalog_index(base)
    resolved = []
    for item in preview.get("rows", []):
        if not item.get("valid"):
            continue
        values = dict(item.get("values") or {})
        ticker = values.pop("ticker")
        enabled = values.pop("enabled", None)
        existing = current.get(ticker)
        row = dict(existing or {})
        if not row:
            required = ("name", "country", "city", "lat", "lon", "seg")
            if any(values.get(field) in (None, "") for field in required):
                row = resolve_ticker(base, ticker, override_name=values.get("name"))
        row.update({key: value for key, value in values.items() if value not in (None, "")})
        row["t"] = ticker
        row["source"] = "import"
        resolved.append({
            "ticker": ticker,
            "enabled": enabled,
            "classification": item.get("classification"),
            "row": row,
        })
    return resolved


def _backup_catalog(base):
    backup_id = datetime.datetime.now().strftime("%Y%m%d_%H%M%S") + "_" + secrets.token_hex(4)
    target = Path(base) / "data" / "backups" / "catalog" / backup_id
    target.mkdir(parents=True, exist_ok=False)
    for name in ("user_watchlist.json", "active_universe.json"):
        source = Path(base) / "data" / name
        if source.exists():
            shutil.copy2(source, target / name)
    return backup_id


def apply_preview(base, preview_id, mode, confirmation, default_companies):
    base = Path(base)
    preview = load_preview(base, preview_id)
    if preview.get("catalog_hash") != _catalog_hash(base):
        raise ApiError(409, "stale_preview", "Catalog changed after preview; create a new preview")
    mode = str(mode or "merge").strip().lower()
    if mode not in {"merge", "replace"}:
        raise ApiError(400, "invalid_import_mode", "Import mode must be merge or replace")
    if mode == "replace" and str(confirmation or "") != REPLACE_CONFIRMATION:
        raise ApiError(400, "replacement_confirmation_required", "Replacement confirmation phrase is incorrect")

    resolved = _resolved_import_rows(base, preview)
    defaults = _default_mapping(default_companies)
    backup_id = None
    with data_lock(base):
        if mode == "replace":
            backup_id = _backup_catalog(base)
            user = {"version": 1, "replace_base": True, "add": {}, "remove": []}
            for item in resolved:
                if item["enabled"] is False:
                    continue
                user["add"][item["ticker"]] = item["row"]
        else:
            user = load_user_watchlist(base)
            add = user.setdefault("add", {})
            remove = {str(t).upper() for t in user.get("remove", [])}
            base_tickers = {
                str(row.get("t") or "").upper()
                for row in read_json(base / "data" / "watchlist.json", {}).get("companies", [])
            }
            for item in resolved:
                ticker = item["ticker"]
                if item["enabled"] is False:
                    if ticker in add and ticker not in base_tickers:
                        add.pop(ticker, None)
                    else:
                        remove.add(ticker)
                    continue
                remove.discard(ticker)
                if item["classification"] != "unchanged" or ticker in add:
                    add[ticker] = item["row"]
            user["remove"] = sorted(remove)
        write_user_watchlist(base, user)
        active = rebuild_active_universe(base, defaults)

    (_preview_dir(base) / f"{preview_id}.json").unlink(missing_ok=True)
    return {
        "ok": True,
        "mode": mode,
        "backup_id": backup_id,
        "applied": len(resolved),
        "skipped": sum(1 for row in preview.get("rows", []) if not row.get("valid")),
        "counts": active.get("counts", {}),
    }
