# -*- coding: utf-8 -*-
"""User-scoped interface templates, previews, apply, and exports."""
import csv, datetime as dt, io, json, os, secrets
from pathlib import Path
from openpyxl import Workbook, load_workbook
from admin_common import ApiError, now_iso
from aicm_io import read_json, write_json_atomic
from auth_context import require_permission
from interface_registry import create_records, list_records

MAX_BYTES=10*1024*1024; MAX_ROWS=10000
COLUMNS=("id","name","provider","category","purpose","monitor_mode","method","url","allow_private_target","timeout_seconds","interval_minutes","expected_statuses","max_latency_ms","json_path","expected_value","body_keyword","freshness_json_path","max_age_minutes","auth_type","credential_configured","notes")


def _validate_headers(headers):
    clean=[str(value or "").strip() for value in headers]
    duplicates=sorted({value for value in clean if value and clean.count(value)>1})
    if duplicates:
        raise ApiError(400,"duplicate_headers","Import contains duplicate headers",duplicates)
    return clean


def interface_template_bytes(variant="blank"):
    wb=Workbook(); info=wb.active; info.title="Instructions"; info.append(["AI Compute Chain interface import template"]); sheet=wb.create_sheet("Interfaces"); sheet.append(COLUMNS)
    if variant=="example": sheet.append(["","Example API","Example","provider","Health","http","GET","https://example.com/health",False,8,15,"200","","data.status","ok","ready","data.updated_at",10,"none",False,""])
    ref=wb.create_sheet("Field Reference"); ref.append(["Field","Required"]); [ref.append([x,x in {"name","monitor_mode"}]) for x in COLUMNS]
    out=io.BytesIO(); wb.save(out); return out.getvalue(),f"interface_template_{variant}.xlsx","application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def parse_upload(filename,body):
    if len(body)>MAX_BYTES: raise ApiError(413,"upload_too_large","Upload exceeds 10 MB")
    suffix=Path(filename).suffix.lower(); rows=[]
    if suffix==".csv":
        reader=csv.DictReader(io.StringIO(body.decode("utf-8-sig")))
        _validate_headers(reader.fieldnames or []); rows=list(reader)
    elif suffix==".xlsx":
        wb=load_workbook(io.BytesIO(body),read_only=True,data_only=True); sheet=wb["Interfaces"] if "Interfaces" in wb.sheetnames else wb.active; values=list(sheet.iter_rows(values_only=True)); headers=_validate_headers(values[0]) if values else []; rows=[dict(zip(headers,row)) for row in values[1:]]
    elif suffix==".json":
        payload=json.loads(body); rows=payload.get("interfaces",payload) if isinstance(payload,dict) else payload
    else: raise ApiError(400,"unsupported_format","Use XLSX, CSV, or JSON")
    if len(rows)>MAX_ROWS: raise ApiError(400,"too_many_rows","Import exceeds 10,000 rows")
    output=[]
    for number,row in enumerate(rows,2):
        clean={k:v for k,v in dict(row).items() if k in COLUMNS and v not in (None,"")}; clean["_row"]=number
        for key in ("timeout_seconds","interval_minutes","max_latency_ms","max_age_minutes"):
            if key in clean: clean[key]=int(clean[key])
        if isinstance(clean.get("expected_statuses"),str): clean["expected_statuses"]=[int(x.strip()) for x in clean["expected_statuses"].split(",") if x.strip()]
        output.append(clean)
    return output


def _preview_path(workspace,preview_id): return workspace.path("previews")/f"{preview_id}.json"
def create_preview(workspace,rows,filename):
    require_permission(workspace.auth,"transfer.import"); preview_id=secrets.token_hex(16); directory=workspace.path("previews"); directory.mkdir(parents=True,exist_ok=True); os.chmod(directory,0o700)
    preview={"preview_id":preview_id,"filename":Path(filename).name,"created_at":now_iso(),"workspace_hash":list_records(workspace)["workspace_hash"],"items":rows,"summary":{"rows":len(rows)}}
    write_json_atomic(_preview_path(workspace,preview_id),preview); os.chmod(_preview_path(workspace,preview_id),0o600); return preview
def apply_preview(workspace,preview_id):
    require_permission(workspace.auth,"transfer.import"); path=_preview_path(workspace,str(preview_id))
    if not path.exists(): raise ApiError(404,"preview_not_found","Preview not found in this workspace")
    preview=read_json(path,{})
    try:
        created=dt.datetime.fromisoformat(str(preview.get("created_at") or "").replace("Z","+00:00"))
    except ValueError:
        created=None
    if not created or created <= dt.datetime.now(dt.timezone.utc)-dt.timedelta(minutes=30):
        path.unlink(missing_ok=True)
        raise ApiError(410,"preview_expired","Import preview has expired")
    current=list_records(workspace)["workspace_hash"]
    if preview.get("workspace_hash")!=current: raise ApiError(409,"stale_preview","Workspace changed after preview")
    payloads=[{k:v for k,v in row.items() if not k.startswith("_") and k!="id"} for row in preview.get("items",[])]
    created=create_records(workspace,payloads)
    path.unlink(missing_ok=True); return {"applied":len(created)}


def export_interfaces(workspace,format_name="json"):
    require_permission(workspace.auth,"transfer.export"); rows=list_records(workspace)["items"]
    safe=[{k:v for k,v in row.items() if k in COLUMNS} for row in rows]
    fmt=str(format_name).lower()
    if fmt=="json": return json.dumps({"schema_version":1,"generated_at":now_iso(),"interfaces":safe},ensure_ascii=False,indent=2).encode(),"interfaces.json","application/json"
    if fmt=="csv":
        out=io.StringIO(); writer=csv.DictWriter(out,fieldnames=COLUMNS); writer.writeheader(); writer.writerows([{**row,"expected_statuses":",".join(map(str,row.get("expected_statuses",[])))} for row in safe]); return ("\ufeff"+out.getvalue()).encode(),"interfaces.csv","text/csv"
    if fmt=="xlsx":
        wb=Workbook(); sheet=wb.active; sheet.title="Interfaces"; sheet.append(COLUMNS)
        for row in safe: sheet.append([",".join(map(str,row.get(k,[]))) if isinstance(row.get(k),list) else row.get(k) for k in COLUMNS])
        meta=wb.create_sheet("Metadata"); meta.append(["generated_at",now_iso()]); output=io.BytesIO(); wb.save(output); return output.getvalue(),"interfaces.xlsx","application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    raise ApiError(400,"unsupported_format","Use XLSX, CSV, or JSON")
