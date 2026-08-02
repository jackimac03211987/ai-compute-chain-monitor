import hashlib
import io
import json
import threading
import unittest
import urllib.error
import urllib.request

from openpyxl import Workbook, load_workbook

import app
from admin_jobs import AdminJobManager
from admin_service import test_interfaces as run_interface_tests
from aicm_io import data_lock
from tests.helpers import ProjectFixture


CATALOG_FILES = ("watchlist.json", "user_watchlist.json", "active_universe.json")


class AdminEndToEndTests(unittest.TestCase):
    def setUp(self):
        self.fx = ProjectFixture()
        self.token = "e2e-token-must-never-appear-in-audit"
        self.companies = [
            {
                "t": "NVDA", "name": "NVIDIA", "country": "美国", "city": "Santa Clara",
                "lat": 37.35, "lon": -121.95, "seg": "GPU", "chain": "AI accelerators",
                "chain_key": "gpu", "source": "base",
            },
            {
                "t": "AMD", "name": "AMD", "country": "美国", "city": "Santa Clara",
                "lat": 37.38, "lon": -121.97, "seg": "GPU", "chain": "AI accelerators",
                "chain_key": "gpu", "source": "base",
            },
        ]
        self.fx.write_json("watchlist.json", {"companies": self.companies})
        self.fx.write_json("user_watchlist.json", {"version": 1, "replace_base": False, "add": {}, "remove": []})
        self.fx.write_json("active_universe.json", {"companies": self.companies, "counts": {"active": 2, "base": 2}})
        self.fx.write_json("live.json", {
            "asof": "2026-07-11 09:00",
            "items": {
                "NVDA": {"t": "NVDA", "p": 160, "chg": 1.2, "source": "yahoo_chart"},
                "AMD": {"t": "AMD", "p": 145, "chg": -0.4, "source": "yahoo_chart"},
            },
        })
        self.fx.write_json("live_status.json", {
            "status": "success", "last_success": "2026-07-11 09:00",
            "fresh_count": 2, "expected_count": 2,
        })
        self.fx.write_json("semi_market.json", {
            "companies": self.companies, "dates": ["2026-07-11"],
            "meta": {"fetched": "2026-07-11 09:00"},
        })
        (self.fx.data / "admin_token.txt").write_text(self.token + "\n", encoding="utf-8")
        (self.fx.base / "index.html").write_text("AI COMPUTE CHAIN PULSE", encoding="utf-8")
        (self.fx.base / "admin").mkdir()
        (self.fx.base / "admin" / "index.html").write_text("AI 算力链运营控制台", encoding="utf-8")
        (self.fx.base / "private").mkdir()
        (self.fx.base / "private" / "index.html").write_text("私有工作空间", encoding="utf-8")

        self.old_globals = (app.BASE, app.DATA, app.ADMIN_TOKEN_PATH, app.default_companies)
        app.BASE = self.fx.base
        app.DATA = self.fx.data
        app.ADMIN_TOKEN_PATH = self.fx.data / "admin_token.txt"
        app.default_companies = lambda: [dict(row) for row in self.companies]
        app.reset_admin_runtime()
        self.server = app.FastThreadingHTTPServer(("127.0.0.1", 0), app.AppHandler)
        self.thread = threading.Thread(target=self.server.serve_forever, daemon=True)
        self.thread.start()
        self.base_url = f"http://127.0.0.1:{self.server.server_port}"

    def tearDown(self):
        self.server.shutdown()
        self.server.server_close()
        self.thread.join(timeout=2)
        app.BASE, app.DATA, app.ADMIN_TOKEN_PATH, app.default_companies = self.old_globals
        app.reset_admin_runtime()
        self.fx.close()

    def request(self, method, path, body=None, token=True, content_type="application/json"):
        if body is None:
            raw = None
        elif isinstance(body, bytes):
            raw = body
        else:
            raw = json.dumps(body).encode("utf-8")
        headers = {"Content-Type": content_type}
        if token:
            headers["X-AICM-Admin-Token"] = self.token
        request = urllib.request.Request(self.base_url + path, data=raw, headers=headers, method=method)
        try:
            response = urllib.request.urlopen(request, timeout=8)
            return response.status, response.headers, response.read()
        except urllib.error.HTTPError as error:
            return error.code, error.headers, error.read()

    def request_json(self, method, path, body=None, token=True):
        status, headers, raw = self.request(method, path, body, token)
        return status, headers, json.loads(raw)

    def catalog_bytes(self):
        return {name: (self.fx.data / name).read_bytes() for name in CATALOG_FILES}

    def catalog_hashes(self):
        return {name: hashlib.sha256(body).hexdigest() for name, body in self.catalog_bytes().items()}

    def test_admin_contracts_templates_and_non_mutating_interface_test(self):
        status, _, payload = self.request_json("GET", "/api/admin/overview", token=False)
        self.assertEqual(status, 403)
        self.assertEqual(payload["error"]["code"], "admin_token_required")

        status, _, payload = self.request_json("GET", "/api/admin/overview")
        self.assertEqual(status, 200)
        self.assertEqual(payload["data"]["catalog"]["active"], 2)

        status, _, payload = self.request_json("GET", "/api/admin/interfaces")
        self.assertEqual(status, 200)
        self.assertEqual(len(payload["data"]["items"]), 8)

        before = self.catalog_hashes()
        result = run_interface_tests(self.fx.base, interface_ids=["catalog"])
        self.assertEqual([item["id"] for item in result["items"]], ["catalog"])
        self.assertEqual(self.catalog_hashes(), before)

        status, headers, raw = self.request("GET", "/api/admin/templates/companies.xlsx?variant=blank")
        self.assertEqual(status, 200)
        self.assertIn("spreadsheetml", headers.get("Content-Type"))
        workbook = load_workbook(io.BytesIO(raw), read_only=True)
        self.assertEqual(workbook.sheetnames, ["Instructions", "Companies", "Field Reference"])

    def test_health_checker_validates_admin_without_echoing_token(self):
        from health_check import check_admin

        result = check_admin(self.base_url, self.token)
        self.assertEqual(result["errors"], [])
        self.assertTrue(result["admin_static_available"])
        self.assertTrue(result["admin_unauthorized_forbidden"])
        self.assertEqual(result["admin_interface_count"], 8)
        self.assertTrue(result["private_static_available"])
        self.assertTrue(result["private_unauthorized_forbidden"])
        self.assertNotIn(self.token, json.dumps(result, ensure_ascii=False))

    def test_preview_merge_export_remove_and_exact_catalog_restore(self):
        original_bytes = self.catalog_bytes()
        original_hashes = self.catalog_hashes()

        csv_body = (
            "ticker,name,country,city,lat,lon,seg,chain,chain_key,enabled\n"
            "NVDA,,,,,,,,,true\n"
            "AICMTEST,Fixture Corp,美国,Austin,30.2672,-97.7431,Testing,Test chain,test,true\n"
        ).encode("utf-8")
        status, _, preview_payload = self.request_json(
            "POST", "/api/admin/import/preview?filename=companies.csv", csv_body
        )
        self.assertEqual(status, 200)
        self.assertEqual(self.catalog_hashes(), original_hashes)

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Companies"
        sheet.append(["ticker", "name"])
        sheet.append(["NVDA", "NVIDIA"])
        xlsx = io.BytesIO()
        workbook.save(xlsx)
        status, _, _ = self.request_json(
            "POST", "/api/admin/import/preview?filename=companies.xlsx", xlsx.getvalue()
        )
        self.assertEqual(status, 200)
        self.assertEqual(self.catalog_hashes(), original_hashes)

        preview_id = preview_payload["data"]["preview_id"]
        status, _, applied = self.request_json(
            "POST", "/api/admin/import/apply", {"preview_id": preview_id, "mode": "merge"}
        )
        self.assertEqual(status, 200)
        self.assertGreaterEqual(applied["data"]["applied"], 1)
        active = self.fx.read_json("active_universe.json")["companies"]
        by_ticker = {row["t"]: row for row in active}
        self.assertEqual(by_ticker["NVDA"]["city"], "Santa Clara")
        self.assertIn("AICMTEST", by_ticker)

        status, _, exported = self.request("GET", "/api/admin/export?format=json&scope=active")
        self.assertEqual(status, 200)
        export_payload = json.loads(exported)
        self.assertEqual(len(export_payload["companies"]), len(active))

        status, _, _ = self.request_json(
            "POST", "/api/admin/companies/toggle", {"ticker": "AICMTEST", "enabled": False}
        )
        self.assertEqual(status, 200)

        with data_lock(self.fx.base):
            for name, body in original_bytes.items():
                (self.fx.data / name).write_bytes(body)
        self.assertEqual(self.catalog_hashes(), original_hashes)

        status, _, audit = self.request_json("GET", "/api/admin/audit")
        self.assertEqual(status, 200)
        self.assertNotIn(self.token, json.dumps(audit, ensure_ascii=False))
        self.assertNotIn(self.token, (self.fx.data / "admin_audit.jsonl").read_text(encoding="utf-8"))

    def test_duplicate_live_and_history_jobs_return_existing_job(self):
        held = []
        manager = AdminJobManager(self.fx.base, runner=lambda job, done: held.append((job, done)))
        app._ADMIN_JOB_MANAGERS[str(self.fx.base.resolve())] = manager
        for kind in ("live", "history"):
            status, _, first = self.request_json("POST", "/api/admin/jobs/start", {"kind": kind})
            self.assertEqual(status, 202)
            status, _, second = self.request_json("POST", "/api/admin/jobs/start", {"kind": kind})
            self.assertEqual(status, 202)
            self.assertEqual(second["data"]["job_id"], first["data"]["job_id"])
            self.assertTrue(second["data"]["already_running"])
        self.assertEqual(len(held), 2)


if __name__ == "__main__":
    unittest.main()
