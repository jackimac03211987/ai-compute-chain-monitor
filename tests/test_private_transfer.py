import io, json, tempfile, unittest
from unittest.mock import patch
from pathlib import Path
from openpyxl import load_workbook
from identity_store import AuthContext
from workspace import WorkspaceContext


def ws(base, tenant, user):
    auth=AuthContext(tenant,user,"operator",frozenset({"interfaces.read","interfaces.write","transfer.import","transfer.export","watchlist.read","watchlist.write"}),"t")
    return WorkspaceContext(base,auth)


class PrivateTransferTests(unittest.TestCase):
    def setUp(self): self.tmp=tempfile.TemporaryDirectory(); self.base=Path(self.tmp.name); self.first=ws(self.base,"tenant_a","user_a"); self.second=ws(self.base,"tenant_b","user_b")
    def tearDown(self): self.tmp.cleanup()

    def test_interface_template_and_secret_free_exports(self):
        from interface_transfer import export_interfaces, interface_template_bytes
        from interface_registry import create_record
        body,_,_=interface_template_bytes("blank")
        self.assertEqual(load_workbook(io.BytesIO(body),read_only=True).sheetnames,["Instructions","Interfaces","Field Reference"])
        create_record(self.first,{"name":"API","provider":"X","category":"provider","purpose":"Test","monitor_mode":"http","method":"GET","url":"https://example.com","interval_minutes":15,"timeout_seconds":8,"expected_statuses":[200],"credential_configured":True})
        for fmt in ("json","csv","xlsx"):
            payload,_,_=export_interfaces(self.first,fmt)
            self.assertNotIn(b"secret-value",payload)

    def test_preview_is_user_bound_and_applies_safe_merge(self):
        from admin_common import ApiError
        from interface_transfer import apply_preview, create_preview, parse_upload
        rows=parse_upload("interfaces.csv",b"name,provider,category,purpose,monitor_mode,method,url,timeout_seconds,interval_minutes\nExample,Vendor,provider,Health,http,GET,https://example.com,8,15\n")
        preview=create_preview(self.first,rows,"interfaces.csv")
        with self.assertRaises(ApiError): apply_preview(self.second,preview["preview_id"])
        result=apply_preview(self.first,preview["preview_id"])
        self.assertEqual(result["applied"],1)

    def test_personal_watchlist_csv_import_is_private(self):
        from personal_watchlist_transfer import import_watchlist
        from personal_watchlist import list_personal_watchlist
        import_watchlist(self.first,"watchlist.csv",b"ticker,name\nNVDA,NVIDIA\n")
        self.assertEqual(list_personal_watchlist(self.first)["count"],1)
        self.assertEqual(list_personal_watchlist(self.second)["count"],0)

    def test_interface_preview_expires_after_thirty_minutes(self):
        from admin_common import ApiError
        from interface_transfer import apply_preview, create_preview
        with patch("interface_transfer.now_iso", return_value="2026-01-01T00:00:00+00:00"):
            preview=create_preview(self.first,[{"name":"Old","monitor_mode":"local_task"}],"old.json")
        with self.assertRaises(ApiError) as error:
            apply_preview(self.first,preview["preview_id"])
        self.assertEqual(error.exception.code,"preview_expired")

    def test_duplicate_import_headers_are_rejected(self):
        from admin_common import ApiError
        from interface_transfer import parse_upload
        with self.assertRaises(ApiError) as error:
            parse_upload("interfaces.csv",b"name,name,monitor_mode\nOne,Two,local_task\n")
        self.assertEqual(error.exception.code,"duplicate_headers")


if __name__=="__main__": unittest.main()
