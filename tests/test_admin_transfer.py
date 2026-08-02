import io
import json
import hashlib
import unittest

from openpyxl import load_workbook

from tests.helpers import ProjectFixture


class TransferTests(unittest.TestCase):
    def setUp(self):
        self.fx = ProjectFixture()
        company = {
            "t": "NVDA",
            "name": "NVIDIA 英伟达",
            "country": "美国",
            "city": "Santa Clara",
            "lat": 37.35,
            "lon": -121.95,
            "seg": "GPU/AI芯片",
        }
        self.fx.write_json("watchlist.json", {"companies": [company]})
        self.fx.write_json(
            "active_universe.json",
            {"companies": [{**company, "source": "base"}], "counts": {"active": 1}},
        )
        self.fx.write_json(
            "user_watchlist.json",
            {"version": 1, "replace_base": False, "add": {}, "remove": []},
        )
        self.fx.write_json(
            "live.json",
            {
                "asof": "2026-07-11 09:00",
                "items": {
                    "NVDA": {
                        "t": "NVDA",
                        "p": 160.5,
                        "chg": 1.25,
                        "market_time": 1783731000,
                        "source": "yahoo_chart",
                    }
                },
            },
        )

    def tearDown(self):
        self.fx.close()

    def test_blank_xlsx_template_has_three_named_sheets(self):
        from admin_transfer import template_bytes

        body, filename, mime = template_bytes("blank")
        workbook = load_workbook(io.BytesIO(body), read_only=True)
        self.assertEqual(
            workbook.sheetnames,
            ["Instructions", "Companies", "Field Reference"],
        )
        self.assertTrue(filename.endswith(".xlsx"))
        self.assertEqual(
            mime,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def test_example_template_contains_company_rows(self):
        from admin_transfer import template_bytes

        body, _, _ = template_bytes("example")
        workbook = load_workbook(io.BytesIO(body), read_only=True, data_only=True)
        rows = list(workbook["Companies"].iter_rows(values_only=True))
        self.assertEqual(rows[0][0], "ticker")
        self.assertEqual(rows[1][0], "NVDA")

    def test_csv_parser_accepts_utf8_bom(self):
        from admin_transfer import parse_upload

        rows = parse_upload(
            "companies.csv",
            "ticker,name,country\nNVDA,NVIDIA 英伟达,美国\n".encode("utf-8-sig"),
        )
        self.assertEqual(rows[0]["ticker"], "NVDA")
        self.assertEqual(rows[0]["name"], "NVIDIA 英伟达")
        self.assertEqual(rows[0]["_row"], 2)

    def test_parser_rejects_unsupported_extension(self):
        from admin_common import ApiError
        from admin_transfer import parse_upload

        with self.assertRaises(ApiError) as raised:
            parse_upload("companies.xlsm", b"not-a-workbook")
        self.assertEqual(raised.exception.code, "unsupported_file_type")

    def test_json_export_has_schema_quotes_and_one_row(self):
        from admin_transfer import export_bytes

        body, filename, mime = export_bytes(
            self.fx.base,
            "json",
            "active",
            include_quotes=True,
        )
        payload = json.loads(body)
        self.assertEqual(payload["schema_version"], 1)
        self.assertEqual(len(payload["companies"]), 1)
        self.assertEqual(payload["companies"][0]["latest_price"], 160.5)
        self.assertTrue(filename.endswith(".json"))
        self.assertEqual(mime, "application/json; charset=utf-8")

    def test_xlsx_export_has_companies_and_metadata_sheets(self):
        from admin_transfer import export_bytes

        body, _, _ = export_bytes(self.fx.base, "xlsx", "base")
        workbook = load_workbook(io.BytesIO(body), read_only=True, data_only=True)
        self.assertEqual(workbook.sheetnames, ["Companies", "Metadata"])
        rows = list(workbook["Companies"].iter_rows(values_only=True))
        self.assertEqual(rows[1][0], "NVDA")

    def catalog_hashes(self):
        output = {}
        for name in ("watchlist.json", "user_watchlist.json", "active_universe.json"):
            output[name] = hashlib.sha256((self.fx.data / name).read_bytes()).hexdigest()
        return output

    def default_companies(self):
        return {
            "NVDA": (
                "NVIDIA 英伟达",
                "美国",
                "Santa Clara",
                37.35,
                -121.95,
                "GPU/AI芯片",
            )
        }

    def test_preview_classifies_rows_without_writing_catalog(self):
        from admin_transfer import create_preview

        before = self.catalog_hashes()
        preview = create_preview(
            self.fx.base,
            [
                {"_row": 2, "ticker": "NVDA", "name": "NVIDIA 英伟达"},
                {"_row": 3, "ticker": "TSLA", "name": "Tesla"},
                {"_row": 4, "ticker": "BAD$CODE", "name": "Bad"},
            ],
            "companies.csv",
        )
        self.assertEqual(preview["summary"]["unchanged"], 1)
        self.assertEqual(preview["summary"]["new"], 1)
        self.assertEqual(preview["summary"]["invalid"], 1)
        self.assertEqual(self.catalog_hashes(), before)

    def test_safe_merge_does_not_erase_existing_fields_with_blanks(self):
        from admin_transfer import apply_preview, create_preview

        preview = create_preview(
            self.fx.base,
            [{"_row": 2, "ticker": "NVDA", "name": "", "city": ""}],
            "companies.csv",
        )
        result = apply_preview(
            self.fx.base,
            preview["preview_id"],
            "merge",
            "",
            self.default_companies,
        )
        self.assertTrue(result["ok"])
        row = self.fx.read_json("active_universe.json")["companies"][0]
        self.assertEqual(row["name"], "NVIDIA 英伟达")
        self.assertEqual(row["city"], "Santa Clara")

    def test_stale_preview_is_rejected(self):
        from admin_common import ApiError
        from admin_transfer import apply_preview, create_preview

        preview = create_preview(
            self.fx.base,
            [{"_row": 2, "ticker": "NVDA", "city": "San Jose"}],
            "companies.csv",
        )
        self.fx.write_json(
            "user_watchlist.json",
            {"version": 1, "replace_base": False, "add": {}, "remove": ["NVDA"]},
        )
        with self.assertRaises(ApiError) as raised:
            apply_preview(
                self.fx.base,
                preview["preview_id"],
                "merge",
                "",
                self.default_companies,
            )
        self.assertEqual(raised.exception.code, "stale_preview")

    def test_expired_preview_reports_expired_not_missing(self):
        from admin_common import ApiError
        from admin_transfer import create_preview, load_preview

        preview = create_preview(
            self.fx.base,
            [{"_row": 2, "ticker": "NVDA"}],
            "companies.csv",
        )
        path = (
            self.fx.data
            / "admin_previews"
            / f'{preview["preview_id"]}.json'
        )
        payload = json.loads(path.read_text(encoding="utf-8"))
        payload["expires_epoch"] = 1
        path.write_text(json.dumps(payload), encoding="utf-8")
        with self.assertRaises(ApiError) as raised:
            load_preview(self.fx.base, preview["preview_id"])
        self.assertEqual(raised.exception.code, "preview_expired")

    def test_replacement_requires_confirmation_and_creates_backup(self):
        from admin_common import ApiError
        from admin_transfer import apply_preview, create_preview

        preview = create_preview(
            self.fx.base,
            [{"_row": 2, "ticker": "NVDA", "name": "NVIDIA 英伟达"}],
            "companies.csv",
        )
        with self.assertRaises(ApiError):
            apply_preview(
                self.fx.base,
                preview["preview_id"],
                "replace",
                "wrong",
                self.default_companies,
            )
        result = apply_preview(
            self.fx.base,
            preview["preview_id"],
            "replace",
            "REPLACE ACTIVE CATALOG",
            self.default_companies,
        )
        self.assertTrue(result["backup_id"])
        backup = self.fx.data / "backups" / "catalog" / result["backup_id"]
        self.assertTrue((backup / "user_watchlist.json").exists())
        self.assertTrue((backup / "active_universe.json").exists())
