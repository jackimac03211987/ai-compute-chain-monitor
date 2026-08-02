# -*- coding: utf-8 -*-
"""Simple private watchlist imports."""
import csv, io, json
from pathlib import Path
from openpyxl import load_workbook
from admin_common import ApiError
from auth_context import require_permission
from personal_watchlist import add_personal_tickers


def import_watchlist(workspace,filename,body):
    require_permission(workspace.auth,"transfer.import"); suffix=Path(filename).suffix.lower()
    if len(body)>10*1024*1024: raise ApiError(413,"upload_too_large","Upload exceeds 10 MB")
    if suffix==".csv": rows=list(csv.DictReader(io.StringIO(body.decode("utf-8-sig"))))
    elif suffix==".json":
        payload=json.loads(body); rows=payload.get("items",payload) if isinstance(payload,dict) else payload
    elif suffix==".xlsx":
        sheet=load_workbook(io.BytesIO(body),read_only=True,data_only=True).active; values=list(sheet.iter_rows(values_only=True)); headers=values[0]; rows=[dict(zip(headers,row)) for row in values[1:]]
    else: raise ApiError(400,"unsupported_format","Use XLSX, CSV, or JSON")
    if len(rows)>10000: raise ApiError(400,"too_many_rows","Import exceeds 10,000 rows")
    return add_personal_tickers(workspace,rows)
